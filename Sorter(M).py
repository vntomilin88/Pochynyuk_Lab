#! /usr/bin/env python3

import os
import re
import time
import openpyxl
import threading
import tkinter as tk
from tkinter import *
from tkinter import ttk
from datetime import date
from datetime import timedelta
from openpyxl.styles import Font, Alignment

def dob(t):
    return time.strftime(t, time.localtime())
    
def nextm(text):
    nextn = int(re.compile(r'(\S+)[MU](\d+)').search(text).group(2))+1
    return re.compile(r'(\S+)[MU](\d+)').search(text).group(1) + 'M' + str(nextn)

def name(sheet):
    wb = openpyxl.load_workbook('Mice.xlsx')
    name1 = nextm(wb[sheet]['B2'].value)
    namevar.set(name1)
    
def click():
    wbu = openpyxl.load_workbook('Urine(M).xlsx')
    sheet = genvar.get()
    nameu = nextm(wbu[sheet]['A2'].value)

    def mainthread(args): #когда подаешь функции лист, нужно задавать не *args а args
        
        def cell(cell, var):
            wb[sheet][args[cell]] = args[var]
            wb[sheet][args[cell]].font = Font(name='Arial', size=12)
            wb[sheet][args[cell]].alignment = Alignment(horizontal="center", vertical="center")
        
        wb = openpyxl.load_workbook(args[0])
        
        wb[sheet].insert_rows(2)
        
        for i in range(1, 28, 2):
            cell(i, i+1)
              
        wb.save(args[0])
        
        #time.sleep(10)
            
    def baset(wb, nm, cm): #base table
        return [wb, 
            'B2', nm,
            'C2', sexvar.get(),
            'D2', genvar.get(),
            'E2', dobvar.get(),
            'F2', harvar.get(),
            'G2', dietvar.get(),
            'H2', watervar.get(),
            'I2', cm]
        
    def addt(j2, k2, l2, m2, n2, o2): #additional table
        return [
            'J2', j2,
            'K2', k2,
            'L2', l2,
            'M2', m2,
            'N2', n2,
            'O2', o2]
    
    def urinet():
        return ['Urine(M).xlsx', 
            'A2', nameu,
            'B2', namevar.get(),
            'C2', sexvar.get(),
            'D2', genvar.get(),
            'E2', dobvar.get(),
            'F2', harvar.get(),
            'G2', dietvar.get(),
            'H2', '',
            'I2', watervar.get(),
            'J2', '',
            'K2', comvar.get(),
            'L2', coluvar.get(),
            'M2', '1',
            'N2', datevar.get(),]
        
    def thread(lists):
        threading.Thread(target=mainthread(lists)).start()
        
    if logvar.get() == 1:
        thread(baset('Mice.xlsx', namevar.get(), comvar.get()) + addt(wvar.get(), kwvar.get(), nul, nul, nul, nul))
    
    if servar.get() == 1:
       thread(baset('Serum(M).xlsx', namevar.get(), comvar.get()) + addt(colvar.get(), slocvar.get(), sercomvar.get(), nul, nul, nul))
            
    if emvar.get() == 1:
        if kidvare.get() == '1':
            thread(baset('Embedded(M).xlsx', namevar.get(), comvar.get()) + addt(nul, elocvar.get(), emcomvar.get(), nul, nul, nul))
        else:
            thread(baset('Embedded(M).xlsx', namevar.get()+'(1)', comvar.get()) + addt(nul, elocvar.get(), emcomvar.get(), nul, nul, nul))
            thread(baset('Embedded(M).xlsx', namevar.get()+'(2)', comvar.get()) + addt(nul, elocvar.get(), emcomvar.get(), nul, nul, nul))
    
    if frovar.get() == 1:
        if kidvarf.get() == '1':
            thread(baset('Frozen(M).xlsx', namevar.get(), comvar.get()) + addt(nul, nul, nul, nul, nul, nul))
        else:
            thread(baset('Frozen(M).xlsx', namevar.get()+'(1)', comvar.get()) + addt(nul, nul, nul, nul, nul, nul))
            thread(baset('Frozen(M).xlsx', namevar.get()+'(2)', comvar.get()) + addt(nul, nul, nul, nul, nul, nul))
            
    if urvar.get() == 1:
        thread(urinet())
        
    if livar.get() == 1:
        thread(baset('Frozen(M).xlsx', namevar.get(), '(Liver) ' + comvar.get()) + addt(nul, nul, nul, nul, nul, nul))
    
    namevar.set(nextm(openpyxl.load_workbook('Mice.xlsx')[genvar.get()]['B2'].value))
    #master.quit()

master = tk.Tk()
master.title('Logger(M)')

#Lists:
genotype = ['WT', 'TRPC3', 'TRPV4', 'ClCK2', 'ClCK2Tg',  'EPAC1', 'EPAC2', 'EPAC12', 'B1-ClCK2']
sex = ['♂', '♀']
diet = ['Regular', 'Na def. (TD.90228)', 'Na def. 0,5%Cl (TD.170568)', '1%NaCl (TD.90229)', '4%NaCl (TD.92034)', '1,6%Na 0,5%Cl (TD.170569)', 'K def. (TD.88239)', '10% KCl (TD.150699)', '6%K 0,5%Cl (TD.150759)', '3%Li2CO3 (TD.150700)', '']
water = ['Regular', '24h WATER DEPRIVATION', 'GSK2193874', '280nM NH4Cl + 0.5% Sucrose', '280nM NaHC03 + 0.5% Sucrose', '']
collection = ['Premortem', 'Postmortem']
collectionu = ['Spot', 'Postmortem', '']
kidneyse = [1, 2]
kidneysf = [1, 2]


fpvar = StringVar(master)
fpvar.set('Y:\\1 - LAB Resources\\Logs\\Mice')
#fpvar.set('F:\\Science\\Scripts\\Science\\Sorter files\\Mice')

sexvar = StringVar(master)
sexvar.set(sex[0])
genvar = StringVar(master)
genvar.set(genotype[0])
datevar = StringVar(master)
date1 = time.strftime(" (%H:%M)", time.localtime())
datevar.set(str(date.today()) + date1)
dobvar = StringVar(master)
date = date.today()-timedelta(weeks=4)
dobvar.set(date)
harvar = StringVar(master)
harvar.set(date.today())
dietvar = StringVar(master)
dietvar.set(diet[0])
watervar = StringVar(master)
watervar.set(water[0])
comvar = StringVar(master)
comvar.set('')
wvar = StringVar(master)
wvar.set('')
kwvar = StringVar(master)
kwvar.set('')
sercomvar = StringVar(master)
sercomvar.set('BD367988 - 1300g(15min)')
slocvar = StringVar(master)
slocvar.set('')
colvar = StringVar(master)
colvar.set(collection[0])
emcomvar = StringVar(master)
emcomvar.set('Cryoprotected (Sucrose)')
elocvar = StringVar(master)
elocvar.set('')
coluvar = StringVar(master)
coluvar.set(collectionu[0])


nulvar = StringVar(master)
nulvar.set('')

logvar = IntVar()
livar = IntVar()
frovar = IntVar()
urvar = IntVar()
servar = IntVar()
emvar = IntVar()
kidvare = StringVar()
kidvare.set(kidneyse[0])
kidvarf = StringVar()
kidvarf.set(kidneysf[0])

nul = nulvar.get()
os.chdir(fpvar.get())
sheet = genvar.get()
wb = openpyxl.load_workbook('Mice.xlsx')
name1 = nextm(wb[sheet]['B2'].value)
namevar = StringVar(master)
namevar.set(name1)


Label(master, text='Folderpath:', font='None 8 bold').grid(row=0, column=0, padx=2, pady=10)
Entry(master, textvariable=fpvar, width=35).grid(row=0, column=1, columnspan=3, padx=2, pady=10)

Label(master, text='Genotype:', font='None 8 bold').grid(row=0, column=4, padx=10)
OptionMenu(master, genvar, *genotype, command=name).grid(row=0, column=5, padx=2, pady=10)

#Main

#ttk.Separator(master, orient="horizontal").grid(row=1, columnspan=10, sticky="we")

Label(master, text='Name', font='None 8 bold').grid(row=2, column=0)
Entry(master, textvariable=namevar, width=8).grid(row=3, column=0, padx=2)

Label(master, text='Sex', font='None 8 bold').grid(row=2, column=1)
OptionMenu(master, sexvar, *sex).grid(row=3, column=1, padx=2)

Label(master, text='DOB',font='None 8 bold').grid(row=2, column=2)
Entry(master, textvariable=dobvar, width=10).grid(row=3, column=2, padx=2)

Label(master, text='Harvested',font='None 8 bold').grid(row=2, column=3)
Entry(master, textvariable=harvar, width=10).grid(row=3, column=3)

Label(master, text='Diet',font='None 8 bold').grid(row=2, column=4)
OptionMenu(master, dietvar, *diet).grid(row=3, column=4)

Label(master, text='Water',font='None 8 bold').grid(row=2, column=5)
OptionMenu(master, watervar, *water).grid(row=3, column=5)

Label(master, text='Main Comments',font='None 8 bold').grid(row=2, column=6, columnspan=2)
Entry(master, textvariable=comvar, width=40).grid(row=3, column=6, columnspan=2, padx=5)

ttk.Separator(master, orient="horizontal").grid(row=4, rowspan=2, columnspan=10, pady=10, sticky="EW")

#log
lrow = 6
Checkbutton(master, text='Log', variable=logvar).grid(row=lrow, column=0, columnspan=2, pady=2, sticky="W")

Label(master, text='Weight (g)').grid(row=lrow, column=2, pady=2)
Entry(master, textvariable=wvar, width=7).grid(row=lrow, column=3, columnspan=1, pady=2)

Label(master, text='Kidneys (g)').grid(row=lrow, column=4, pady=2)
Entry(master, textvariable=kwvar, width=5).grid(row=lrow, column=5, columnspan=1, pady=2)

#Urine
urow = 7
Checkbutton(master, text='Urine', variable=urvar).grid(row=urow, column=0, columnspan=2, pady=2, sticky="W")

Label(master, text='Collection:').grid(row=urow, column=2, pady=2)
OptionMenu(master, coluvar, *collectionu).grid(row=urow, column=3, pady=2)

Label(master, text='Date:').grid(row=urow, column=4, pady=2)
Entry(master, textvariable=datevar, width=17).grid(row=urow, column=5, pady=2)

#Frozen
frow = 8
Checkbutton(master, text='Frozen', variable=frovar).grid(row=frow, column=0, columnspan=2, pady=2, sticky="W")

Label(master, text='Kidney #:').grid(row=frow, column=2, pady=2)
OptionMenu(master, kidvarf, *kidneysf).grid(row=frow, column=3, pady=2)

Checkbutton(master, text='Liver', variable=livar).grid(row=frow, column=4, columnspan=1, pady=2)

#Serum
srow = 9
Checkbutton(master, text='Serum', variable=servar).grid(row=srow, column=0, columnspan=2, pady=2, sticky="W")

Label(master, text='Collection:').grid(row=srow, column=2, pady=2)
OptionMenu(master, colvar, *collection).grid(row=srow, column=3, pady=2)

Label(master, text='Box:').grid(row=srow, column=4, pady=5)
Entry(master, textvariable=slocvar, width=6).grid(row=srow, column=5, pady=2)

Label(master, text='com(Serum):').grid(row=srow, column=6, pady=2)
Entry(master, textvariable=sercomvar, width=23).grid(row=srow, column=7, columnspan=2, pady=2, padx=2)

#Embedded
erow = 10
Checkbutton(master, text='Embedded', variable=emvar).grid(row=erow, column=0, columnspan=2, pady=2, sticky="W")

Label(master, text='Kidney #:').grid(row=erow, column=2, pady=2)
OptionMenu(master, kidvare, *kidneyse).grid(row=erow, column=3, pady=2)

Label(master, text='Box:').grid(row=erow, column=4, pady=2)
Entry(master, textvariable=elocvar, width=6).grid(row=erow, column=5, pady=2)

Label(master, text='com(Embed):').grid(row=erow, column=6, pady=2)
Entry(master, textvariable=emcomvar, width=23).grid(row=erow, column=7, columnspan=2, pady=2, padx=2)


Button(master, text='Submit', width=7, command=click).grid(row=11, column=0, columnspan=2, pady=10)


master.mainloop()
