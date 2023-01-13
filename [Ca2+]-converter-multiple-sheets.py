"""
Calcium converter:

#Rmin = Minimal Ratio value for a single cell trace in ABSENCE OF CALCIUM (5mM EGTA)  
#Rmax = Maximal Ratio value for a single cell trace in PRESENCE OF SATURATING CALCIUM (5mM Ionomycin + 5mM Calcium)
#Kd = dissociastion constant for fura2-Calcium binding, Oleh uses 135
#Sf2 = fluorescence at 380nm wavelength in ABSENCE OF CALCIUM for the most reasonalbe cell trace
#Sb2 = fluorescence at 380nm wavelength in PRESENCE OF SATURATING CALCIUM for the most reasonalbe cell trace
"""

import os
import openpyxl


#os.chdir('C:\\Users\\vntom\\Desktop')
os.chdir('C:\\Users\\vtomilin\\Desktop')

cond = {
        'example_set_name': ['date&cells','Rmin','Rmax','Kd','Sf2','Sb2'],
        '20_7_mpkCCD': ['2020.7.22 mpkCCD EGTA-Iono 009',0.11, 1.9, 225, 1650, 350],
        '22_10_5_mCD': ['2022.10.5', 0.13, 1.4, 225, 1550, 300],
        '22_10_13_mCD': ['2022.10.13 mCD EGTA-Iono 002(6)', 0.135, 1.32, 225, 1900, 400]
        }

Rmin = cond['22_10_13_mCD'][1]
Rmax = cond['22_10_13_mCD'][2]
Kd = 225
Sf2 = cond['22_10_13_mCD'][4]
Sb2 = cond['22_10_13_mCD'][5]

rwb = openpyxl.load_workbook('Ratio.xlsx')
rwb.sheetnames

i = 2
m = 0

sheets = []
for s in rwb:
    sheets.append(s)

l = len(sheets)-1



for sheet in sheets[:l]:
    
    mr = sheet.max_row
    mc = sheet.max_column
    
    #calculation into calcium
    for r in range(2, mr+1):
        for c in range (2, mc+1):
            
            R = sheet.cell(row=r, column=c)
            
            if R.value == None:
                continue
            else:
                R.value = Kd*(Sf2/Sb2)*(R.value - Rmin)/(Rmax-R.value)
    
    #average
    for r in range(2, mr+1):
        n = 0
        summ = 0

        for c in range(2, mc+2):
                        
            if sheet.cell(row=r, column=c).value == None:
                n = n+1
                if mc-n == 0:
                    continue
                else:
                    sheets[-1].cell(row=r, column=i).value = summ/(mc-n)
            else:
                summ = summ + sheet.cell(row=r, column=c).value
                sheets[-1].cell(row=r, column=i).value = summ/(mc-n)

    
    #SD            
    for r in range(2, mr+1):
        n = 0
        sigma = 0
        
        for c in range(2, mc+2):
            
            if sheet.cell(row=r, column=c).value == None:
                n = n+1
                if mc-n == 0:
                    continue
                else:
                    sheets[-1].cell(row=r, column=i+1).value = ((sigma/(mc-n))**0.5)/((mc-n))**0.5
            else:
                sigma = sigma + (sheet.cell(row=r, column=c).value - sheets[-1].cell(row=r, column=i).value)**2
                sheets[-1].cell(row=r, column=i+1).value = ((sigma/(mc-n))**0.5)/((mc-n))**0.5
    
    sheets[-1].cell(row=1, column=i).value = rwb.sheetnames[m]
    sheets[-1].cell(row=1, column=i+1).value = f'SD ({sheet.max_column})'
    
    m+=1
    i+=2

for i in range(6):
    sheets[-1].cell(row=i+2, column=7).value = cond['example_set_name'][i]
    sheets[-1].cell(row=i+2, column=8).value = cond['22_10_13_mCD'][i]
    
rwb.save('Calcium.xlsx')
