# -*- coding: utf-8 -*-
"""
Created on Fri Nov 4 14:06:32 2022

@author: vtomilin

Program that merges separate calcium experiments into one ratio file along marking each set with the experiments
date and name, as well as creating a separate sheet of averages per experiment.
"""

import os
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

sht_nms = ['IC','PC'] #sheet names
os.chdir('E:\\Oblaco\\Science\\Analysis\\2022.9.13 Piezo\\Yoda-1\\V4\\RAW') #curent working directory Windows  
# os.chdir('/Terra/Oblachko/Science/Analysis/2022.9.13 Piezo/Yoda-1/WD/RAW') #curent working directory Linux

colors = ['d4e7e0','dddef1','f5dddd','e8e4d5'] #list of colors chosen at random for each experiment to better distinguish each experimental ratios

Ratio = Workbook() #creating the file with all the collected ratios
Ratio.active.title = 'Averages'
Averages = Ratio.active
Sheet1 = Ratio.create_sheet(title=sht_nms[0])
Sheet2 = Ratio.create_sheet(title=sht_nms[1])


def copy_ratios(aggrgtn_sht,exprmnt_sht,color,file_name,avrg_col_n,coe):
    '''
    Adds the ratios from provided sheet to the matching sheet of the final ratio file:

    aggrgtn_sht (aggregation_sheet) - the final sheet we are creating to aggregating ratios from same type of experimental 
    result sheets

    exprmnt_sht (experiment_sheet) - one sheet from the file containing results of a single experiment that is to be copied
    into the coresponding aggregation sheet

    color - color that is used to mark the single set of experimental ratios in the aggretating file

    file_name - the individual experimental file that the traces are being lifted out of

    avrg_col_n - the number needed to determine the column placement (in the Ratio.xlsx file [Averages] Tab) for the averages of 
    experimental traces for each individual experiment. I decided to use the os.listdir(), which creates a list from the files in
    the folder being processed, index of each file corresponds to the average column, just need to add 2 to place them in the 
    second column of the [Averages] sheet of the final ratio file to start with.
    
    coe - coefficinent needed to separate the first sheet average columns being aggregated (IC) from second (PC) since same 
    function is used on both sheets and we do not want the average values to be overwriten by average values of the second sheet
    '''
    if aggrgtn_sht.max_column==1:
        lcol = 0 #(lcol - last_column) aggrgtn_sht.max_column function on an empty final file gives value=1, in order to avoid starting copying to the 2nd column, have to make an exception
        start = 1 #we want the first column in the final ratio file to be the time, but there is no need to add time for every set of ratios from an experiment
    else:
        lcol = aggrgtn_sht.max_column #takes the number of the last column and adds one to create a spacer
        while aggrgtn_sht.cell(column=(lcol), row=1).value == None: #here max_column can also give the wrong number and create large space between the experiments, to prevent this we substract until we reach a cell that is filled with a value
            lcol -=1
        start = 2 #this makes sure time is avoided for all the experiments after the first
        
    for row in range(1,exprmnt_sht.max_row+1):
        num=0
        summ=0
        #Main destination sheet
        for col in range(start,exprmnt_sht.max_column+1):
            current_cell = aggrgtn_sht.cell(column=(lcol)+col, row=row) #current targeted cell in the final aggregated sheet
            

            if exprmnt_sht.cell(column=col, row=row).value == None: #sometimes max_column counts empty cells, to avoid wide empty spaces inplemented this contion
                pass
            elif row==1 and exprmnt_sht.cell(column=col, row=row).value[0]=='#': #making a carveout for the Time and first row of trace numbers
                current_cell.value = exprmnt_sht.cell(column=col, row=row).value[:4]+f'({file_name[:-5]})' #modifying the first row not to add experiment from whitch it comes to the trace number
                current_cell.fill = PatternFill('solid', fgColor=color) #color the cell for better dissernment in the final list
                num +=1
            else:
                current_cell.fill = PatternFill('solid', fgColor=color) #color the cell for better dissernment in the final list
                current_cell.value = exprmnt_sht.cell(column=col, row=row).value #copy the
                if col==1 or exprmnt_sht.cell(column=col, row=row).value==None:
                    continue
                else:
                    summ +=exprmnt_sht.cell(column=col, row=row).value
                    num +=1
        
        #Averages sheet
        if row==1: #The title row
            Averages.cell(column=avrg_col_n.index(file_name)+coe, row=row).value = f'{file_name[:-5]} - {num} - {aggrgtn_sht.title}'
            Averages.cell(column=avrg_col_n.index(file_name)+coe, row=row).fill = PatternFill('solid', fgColor=color)
        else:        
            Averages.cell(column=avrg_col_n.index(file_name)+coe, row=row).value = summ/(num+0.00001) #aggrgtn_sht.title #.{aggrgtn_sht.cell(column=lcol+2, row=row).coordinate}:{aggrgtn_sht.cell(column=lcol+sheet.max_column, row=row).coordinate})' #creates an average of each individual tubule 
            Averages.cell(column=avrg_col_n.index(file_name)+coe, row=row).fill = PatternFill('solid', fgColor=color)

try:
    os.remove('Ratio.xlsx') #atempts to remove the previous calculations for Ratio
except:
    pass


dir_files = os.listdir()
dir_files.remove('Calcium.xlsx')

colors_used = ['start'] #first value to be compared to will not be a color, just to get the loop going

for file in sorted(dir_files): #reads all the individual experiment ratio files and loads them for import into the final ratio file
    ccolor = random.choice(colors) #current color
    print(file)
    if ccolor == colors_used[-1]: #in case we get a random match with the previous color used, we will use a color outside of the colors list
        ccolor = 'DDDDDD'
        copy_ratios(aggrgtn_sht=Sheet1, exprmnt_sht=load_workbook(file)[sht_nms[0]], color=ccolor, file_name=file, avrg_col_n=dir_files, coe=2)
        copy_ratios(aggrgtn_sht=Sheet2, exprmnt_sht=load_workbook(file)[sht_nms[1]], color=ccolor, file_name=file, avrg_col_n=dir_files, coe=3+len(dir_files))
    else:
        copy_ratios(aggrgtn_sht=Sheet1, exprmnt_sht=load_workbook(file)[sht_nms[0]], color=ccolor, file_name=file, avrg_col_n=dir_files, coe=2)
        copy_ratios(aggrgtn_sht=Sheet2, exprmnt_sht=load_workbook(file)[sht_nms[1]], color=ccolor, file_name=file, avrg_col_n=dir_files, coe=3+len(dir_files))
    
    colors_used.append(ccolor)


Ratio.save('Ratio.xlsx')