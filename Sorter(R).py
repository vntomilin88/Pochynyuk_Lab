
import os
import re
import time
import openpyxl
import threading
from tkinter import *
from openpyxl.styles import Font, Alignment

def dob(t):
    return time.strftime(t, time.localtime())
    
def nextm(text):
    nextn = int(re.compile(r'(\w+)(\d+)').search(text).group(2))+1
    return re.compile(r'(\w+)(\d+)').search(text).group(1) + str(nextn)

def click():
    nul = nulvar.get()
    os.chdir(fpvar.get())
    sheet = genvar.get()
    wb = openpyxl.load_workbook('Rats.xlsx')
    wbu = openpyxl.load_workbook('Urine(R).xlsx')
    name = nextm(wb[sheet]['B2'].value)
    nameu = nextm(wbu[sheet]['A2'].value)
    
    def mainthread(args): #когда подаешь функции лист, нужно задавать не *args а args
        
        def cell(cell, var):
            wb[sheet][args[cell]] = args[var]
            wb[sheet][args[cell]].font = Font(size=12)
            wb[sheet][args[cell]].alignment = Alignment(horizontal="center", vertical="center")
        
        wb = openpyxl.load_workbook(args[0])
        
        wb[sheet].insert_rows(2)
        
        for i in range(1, 28, 2):
            cell(i, i+1)
              
        wb.save(args[0])
    
    def baset(wb, nm, cm): #base table
        return [wb, 
            'B2', nm,
            'C2', sexvar.get(),
            'D2', genvar.get(),
            'E2', dobvar.get(),
            'F2', harvar.get(),
            'G2', dietvar.get(),
            'H2', watervar.get(),
            'I2', cm]
        
    def addt(j2, k2, l2, m2, n2, o2): #additional table
        return [
            'J2', j2,
            'K2', k2,
            'L2', l2,
            'M2', m2,
            'N2', n2,
            'O2', o2]
    
    def urinet():
        return ['Urine(R).xlsx', 
            'A2', nameu,
            'B2', name,
            'C2', sexvar.get(),
            'D2', genvar.get(),
            'E2', dobvar.get(),
            'F2', harvar.get(),
            'G2', dietvar.get(),
            'H2', '',
            'I2', watervar.get(),
            'J2', '',
            'K2', comvar.get(),
            'L2', coluvar.get(),
            'M2', '1',
            'N2', datevar.get(),]
        
    def thread(lists):
        threading.Thread(target=mainthread(lists)).start()
        
    thread(baset('Rats.xlsx', name, comvar.get()) + addt(wvar.get(), kwvar.get(), nul, nul, nul, nul))
    
    if servar.get() == 1:
       thread(baset('Serum(R).xlsx', name, comvar.get()) + addt(colvar.get(), slocvar.get(), sercomvar.get(), nul, nul, nul))
            
    if emvar.get() == 1:
        if kidvare.get() == '1':
            thread(baset('Embedded(R).xlsx', name, comvar.get()) + addt(elocvar.get(), emcomvar.get(), nul, nul, nul, nul))
        else:
            thread(baset('Embedded(R).xlsx', name+'(1)', comvar.get()) + addt(elocvar.get(), emcomvar.get(), nul, nul, nul, nul))
            thread(baset('Embedded(R).xlsx', name+'(2)', comvar.get()) + addt(elocvar.get(), emcomvar.get(), nul, nul, nul, nul))
    
    if frovar.get() == 1:
        if kidvarf.get() == '1':
            thread(baset('Frozen(R).xlsx', name, comvar.get()) + addt(nul, nul, nul, nul, nul, nul))
        else:
            thread(baset('Frozen(R).xlsx', name+'(1)', comvar.get()) + addt(nul, nul, nul, nul, nul, nul))
            thread(baset('Frozen(R).xlsx', name+'(2)', comvar.get()) + addt(nul, nul, nul, nul, nul, nul))
            
    if urvar.get() == 1:
        thread(urinet())
        
    if livar.get() == 1:
        thread(baset('Frozen(R).xlsx', name, '(Liver) ' + comvar.get()) + addt(nul, nul, nul, nul, nul, nul))
       
    #master.quit()
    
master = Tk()
master.title('Logger(R)')

#Lists:
genotype = ['WT', 'PCK453']
sex = ['♂', '♀']
diet = ['Regular', '10% KCl (TD.170699)', '6%K 0,5%Cl (TD.170759)']
water = ['Regular', '24h WATER DEPRIVATION', 'GSK2193874']
collection = ['Premortem', 'Postmortem']
collectionu = ['Spot', 'Postmortem']
kidneyse = [1, 2]
kidneysf = [1, 2]


fpvar = StringVar(master)
#fpvar.set('Y:\\Logs')
fpvar.set('F:\\Science\\Scripts\\Science\\Sorter files')

sexvar = StringVar(master)
sexvar.set(sex[0])
genvar = StringVar(master)
genvar.set(genotype[0])
datevar = StringVar(master)
date1 = time.strftime("%Y/%m/%e (%H:%M)", time.localtime())
datevar.set(time.strftime(date1, time.localtime()))
dobvar = StringVar(master)
date = dob('%Y') + '/' + str(int(dob('%m'))-2) + '/' + dob('%e')
dobvar.set(time.strftime(date, time.localtime()))
harvar = StringVar(master)
harvar.set(time.strftime("%Y/%m/%e", time.localtime()))
dietvar = StringVar(master)
dietvar.set(diet[0])
watervar = StringVar(master)
watervar.set(water[0])
comvar = StringVar(master)
comvar.set('')
wvar = StringVar(master)
wvar.set('')
kwvar = StringVar(master)
kwvar.set('')
sercomvar = StringVar(master)
sercomvar.set('BD367988 - 1300g(15min)')
slocvar = StringVar(master)
slocvar.set('')
colvar = StringVar(master)
colvar.set(collection[0])
emcomvar = StringVar(master)
emcomvar.set('Cryoprotected (Sucrose)')
elocvar = StringVar(master)
elocvar.set('')
coluvar = StringVar(master)
coluvar.set(collectionu[0])

nulvar = StringVar(master)
nulvar.set('')

livar = IntVar()
frovar = IntVar()
urvar = IntVar()
servar = IntVar()
emvar = IntVar()
kidvare = StringVar()
kidvare.set(kidneyse[0])
kidvarf = StringVar()
kidvarf.set(kidneysf[0])

Label(master, text='Folderpath:').grid(row=0, column=0)
Entry(master, textvariable=fpvar, width=40).grid(row=0, column=1, columnspan=3)

#Main
Label(master, text='Sex', font=(None, 8)).grid(row=1, column=0)
OptionMenu(master, sexvar, *sex).grid(row=2, column=0)

Label(master, text='Genotype').grid(row=1, column=1)
OptionMenu(master, genvar, *genotype).grid(row=2, column=1)

Label(master, text='DOB').grid(row=1, column=2)
Entry(master, textvariable=dobvar, width=10).grid(row=2, column=2)

Label(master, text='Harvested').grid(row=1, column=3)
Entry(master, textvariable=harvar, width=10).grid(row=2, column=3)

Label(master, text='Diet').grid(row=1, column=4)
OptionMenu(master, dietvar, *diet).grid(row=2, column=4)

Label(master, text='Water').grid(row=1, column=5)
OptionMenu(master, watervar, *water).grid(row=2, column=5)

Label(master, text='Main Comments').grid(row=1, column=6, columnspan=2)
Entry(master, textvariable=comvar, width=25).grid(row=2, column=6, columnspan=2)

Label(master, text='Weight (g)').grid(row=1, column=8)
Entry(master, textvariable=wvar, width=7).grid(row=2, column=8, columnspan=1)

Label(master, text='Kidneys (g)').grid(row=1, column=9)
Entry(master, textvariable=kwvar, width=5).grid(row=2, column=9, columnspan=1)


#Urine
urow = 3
Checkbutton(master, text='Urine', variable=urvar).grid(row=urow, column=0, columnspan=2)

Label(master, text='Collection:').grid(row=urow, column=2)
OptionMenu(master, coluvar, *collectionu).grid(row=urow, column=3)

Label(master, text='Date:').grid(row=urow, column=4)
Entry(master, textvariable=datevar, width=17).grid(row=urow, column=5)

#Frozen
frow = 4
Checkbutton(master, text='Frozen', variable=frovar).grid(row=frow, column=0, columnspan=2)

Label(master, text='Kidney #:').grid(row=frow, column=2)
OptionMenu(master, kidvarf, *kidneysf).grid(row=frow, column=3)

Checkbutton(master, text='Liver', variable=livar).grid(row=frow, column=4, columnspan=1)

#Serum
srow = 5
Checkbutton(master, text='Serum', variable=servar).grid(row=srow, column=0, columnspan=2)

Label(master, text='Collection:').grid(row=srow, column=2)
OptionMenu(master, colvar, *collection).grid(row=srow, column=3)

Label(master, text='Box:').grid(row=srow, column=4)
Entry(master, textvariable=slocvar, width=6).grid(row=srow, column=5)

Label(master, text='Serum Comm.:').grid(row=srow, column=6)
Entry(master, textvariable=sercomvar, width=23).grid(row=srow, column=7, columnspan=2)

#Embedded
erow = 6
Checkbutton(master, text='Embedded', variable=emvar).grid(row=erow, column=0, columnspan=2)

Label(master, text='Kidney #:').grid(row=erow, column=2)
OptionMenu(master, kidvare, *kidneyse).grid(row=erow, column=3)

Label(master, text='Box:').grid(row=erow, column=4)
Entry(master, textvariable=elocvar, width=6).grid(row=erow, column=5)

Label(master, text='Embed Comm.:').grid(row=erow, column=6)
Entry(master, textvariable=emcomvar, width=23).grid(row=erow, column=7, columnspan=2)





Button(master, text='Submit', width=7, command=click).grid(row=8, column=0, columnspan=2)


master.mainloop()
